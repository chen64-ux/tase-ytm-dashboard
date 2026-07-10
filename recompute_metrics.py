# -*- coding: utf-8 -*-
"""
recompute_metrics.py
מחליף את שלב ה-win32com (פתיחת Excel ברקע לאילוץ חישוב מחדש).
קורא את ytm_computed.xlsx, מחשב בפייתון טהור (bond_math.py) את העמודות
AY (תשואה לפדיון), AZ (פרמיית המרה), BA (Duration), BB (Modified Duration),
BC (טווח לפדיון עד המרה סופית), BD (רגישות) - וכותב אותן כערכים קבועים.

יתרון: לא דורש Windows/Excel מותקן - רץ בכל מקום (כולל Termux בטלפון,
GitHub Actions, שרת ענן).

שימוש:
    python3 recompute_metrics.py ytm_computed.xlsx
    (עורך את הקובץ במקום - מומלץ לגבות לפני הרצה ראשונה)
"""

import sys
from datetime import date, datetime
from bond_math import compute_bond_metrics
import openpyxl


def recompute(xlsx_path: str, settlement: date = None):
    if settlement is None:
        settlement = date.today()

    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    ws = wb["Sheet1"]

    updated = 0
    skipped = 0

    for row in range(2, ws.max_row + 1):
        name = ws[f"A{row}"].value
        if not name:
            continue

        price = ws[f"E{row}"].value          # שער אחרון
        rate_pct = ws[f"AD{row}"].value       # ריבית (%)
        maturity_cell = ws[f"AE{row}"].value  # מועד פדיון
        conv_qty = ws[f"AI{row}"].value       # כמות אג"ח למניה אחת
        conv_price = ws[f"AJ{row}"].value     # שער מניית המרה
        conv_final = ws[f"AK{row}"].value     # מועד המרה סופי

        maturity = None
        if isinstance(maturity_cell, datetime):
            maturity = maturity_cell.date()
        elif isinstance(maturity_cell, date):
            maturity = maturity_cell

        # AY, BA, BB - תשואה, Duration, Modified Duration
        if price is not None and rate_pct is not None and maturity is not None:
            res = compute_bond_metrics(settlement, maturity, rate_pct, price)
            ws[f"AY{row}"] = res["ytm"]
            ws[f"BA{row}"] = res["duration"]
            ws[f"BB{row}"] = res["mduration"]
            if res["ytm"] is not None:
                updated += 1
            else:
                skipped += 1
        else:
            ws[f"AY{row}"] = None
            ws[f"BA{row}"] = None
            ws[f"BB{row}"] = None
            skipped += 1

        # AZ - פרמיית המרה = AI*E/AJ - 1
        if conv_qty is not None and price is not None and conv_price:
            ws[f"AZ{row}"] = conv_qty * price / conv_price - 1
        else:
            ws[f"AZ{row}"] = None

        # BC - טווח לפדיון עד המרה סופית = (AK - today)/365
        if isinstance(conv_final, (datetime, date)):
            cf = conv_final.date() if isinstance(conv_final, datetime) else conv_final
            ws[f"BC{row}"] = (cf - settlement).days / 365
        else:
            ws[f"BC{row}"] = None

        # BD - רגישות = 1%/BA
        ba_val = ws[f"BA{row}"].value
        if ba_val:
            ws[f"BD{row}"] = 0.01 / ba_val
        else:
            ws[f"BD{row}"] = None

    wb.save(xlsx_path)
    print(f"[{datetime.now():%Y-%m-%d %H:%M:%S}] חושבו {updated} אג\"ח, {skipped} דולגו (חסרים נתונים) -> {xlsx_path}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("שימוש: python3 recompute_metrics.py <path_to_ytm_computed.xlsx>")
        sys.exit(1)
    recompute(sys.argv[1])
