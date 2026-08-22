# -*- coding: utf-8 -*-
"""
convert_weekly_review.py
קורא את קובץ "סקירה שבועית" (Excel, לשונית יחידה עם 6 סעיפים מופרדים
ע"י שורות כותרת "N. שם הסעיף") ומפיק weekly_review.json - קובץ שהדשבורד
קורא ומציג בלשונית "סקירה שבועית". יש להריץ מחדש ולהעלות מדי שבוע,
כשמתקבל קובץ סקירה חדש (זה לא נשלף אוטומטית - זה תוכן ערוך ידנית).

שימוש:
    python3 convert_weekly_review.py "סקירה שבועית 16-21 אוגוסט 2026.xlsx"
"""

import json
import re
import sys

import openpyxl

SECTION_HEADER_RE = re.compile(r"^\d+\.\s")

# עמודות (0-based) שמכילות לפעמים אחוז גולמי כמספר (לא כמחרוזת "-1.4%"
# מוכנה) - לפי שם הסעיף (התאמה חלקית בכותרת). ערכי float בעמודות אלו
# יומרו לתצוגת אחוזים; ערכי מחרוזת (שכבר מוכנים) יוצגו כפי שהם.
PERCENT_COLUMNS_BY_SECTION_KEYWORD = {
    "מדדי מניות": [3],
    "ארה\"ב": [6],       # דוחות כספיים - ארה"ב: עמודה G (תגובת מניה)
    "ישראל": [5],        # דוחות כספיים - ישראל: עמודה F (תגובת מניה)
}


def percent_cols_for(section_title):
    for keyword, cols in PERCENT_COLUMNS_BY_SECTION_KEYWORD.items():
        if keyword in section_title:
            return cols
    return []


def convert(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    title = rows[0][0] if rows and rows[0][0] else None
    sources_note = rows[1][0] if len(rows) > 1 and rows[1][0] else None

    sections = []
    current = None
    expecting_headers = False

    for row in rows[2:]:
        non_empty = [(i, v) for i, v in enumerate(row) if v is not None and str(v).strip() != ""]
        if not non_empty:
            continue  # שורה ריקה - מפרידה, מדלגים

        first_col_only = len(non_empty) == 1 and non_empty[0][0] == 0
        first_val = str(non_empty[0][1]).strip() if non_empty else ""

        if first_col_only and SECTION_HEADER_RE.match(first_val):
            if current:
                sections.append(current)
            current = {"title": first_val, "headers": None, "rows": [],
                       "percent_cols": percent_cols_for(first_val), "footnotes": []}
            expecting_headers = True
            continue

        if current is None:
            continue  # תוכן לפני הסעיף הראשון (לא אמור לקרות אחרי title/sources_note)

        if expecting_headers:
            current["headers"] = [str(v).strip() if v is not None else "" for v in row[:len(non_empty) + 2]]
            # חותכים עמודות ריקות בסוף
            while current["headers"] and current["headers"][-1] == "":
                current["headers"].pop()
            expecting_headers = False
            continue

        if first_col_only:
            # שורת הערה (משפט ארוך יחיד בעמודה הראשונה בלבד) - שייכת לסעיף הנוכחי
            current["footnotes"].append(first_val)
            continue

        # שורת נתונים רגילה - שומרים את הערכים הגולמיים (מספר/מחרוזת) עד
        # לאורך כותרות הסעיף
        width = len(current["headers"]) if current["headers"] else len(non_empty)
        current["rows"].append(list(row[:width]))

    if current:
        sections.append(current)

    return {"title": title, "sources_note": sources_note, "sections": sections}


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print('שימוש: python3 convert_weekly_review.py "סקירה שבועית.xlsx"')
        sys.exit(1)

    data = convert(sys.argv[1])
    with open("weekly_review.json", "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"✅ הומרו {len(data['sections'])} סעיפים:")
    for s in data["sections"]:
        print(f"  - {s['title']}: {len(s['rows'])} שורות, {len(s['footnotes'])} הערות")
    print("נשמר ל-weekly_review.json")
