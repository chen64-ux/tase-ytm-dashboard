"""
סקריפט יומי מלא: מוריד נתוני מסחר מהבורסה, מעדכן את ytm_computed.xlsx,
מחשב מחדש בפייתון טהור (YTM/פרמייה/Duration - ללא צורך ב-Excel מותקן),
שולף שער דולר עדכני מבנק ישראל, בונה עקום תשואות ממשלתי מתוך אותו קובץ
CSV (אג"ח "ממשל שקלית"), בונה מחדש את ytm_dashboard.html, ומוסיף בראש
הדף חותמת "עודכן לאחרונה" בשעון ישראל.

לא תלוי יותר ב-Windows/Excel - ניתן להריץ בכל מקום: Windows Task
Scheduler, שרת לינוקס, GitHub Actions, או אפילו Termux בטלפון.

*** לפני הרצה ראשונה: לערוך את הגדרות ה-CONFIG למטה ***

דרישות התקנה (חד פעמי):
    pip install requests openpyxl python-dateutil scipy tzdata

קבצים נדרשים באותה תיקייה:
    bond_math.py, recompute_metrics.py, build_dashboard.py

הרצה ידנית לבדיקה:
    python run_daily_update.py
"""

import csv
import datetime
import json
import os
import pathlib
import re
import subprocess
import sys
from zoneinfo import ZoneInfo

import requests
import openpyxl

from recompute_metrics import recompute
from add_new_convertible_bonds import add_new_bonds

# ============================== CONFIG ==============================
# נתיבים יחסיים לתיקיית הריפו - כך שאותו סקריפט עובד גם ב-GitHub Actions
# וגם בהרצה מקומית (Windows/לינוקס/Termux), בלי לשנות כלום.
_REPO_DIR = pathlib.Path(__file__).parent
XLSX_PATH = str(_REPO_DIR / "ytm_computed.xlsx")
DASHBOARD_PATH = str(_REPO_DIR / "docs" / "ytm_dashboard.html")

DOWNLOADS_DIR = _REPO_DIR / "downloads"
LOG_PATH = _REPO_DIR / "run_log.txt"
BUILD_DASHBOARD_SCRIPT = _REPO_DIR / "build_dashboard.py"
# קובץ ייצוא תיק ני"ע מהברוקר - לא מתעדכן אוטומטית, יש להחליף אותו ידנית
# ב-repo (אותו שם קובץ בדיוק) בכל פעם שהתיק משתנה משמעותית.
HOLDINGS_XLSX_PATH = _REPO_DIR / "holdings.xlsx"
# נתוני מכפילים (רווח/הון) מדוח "מבט עומק" תקופתי - לעדכן ידנית אחרי
# כל דוח כספי חדש (ראה ההערה בתוך הקובץ עצמו).
STOCK_FUNDAMENTALS_PATH = _REPO_DIR / "stock_fundamentals.json"
# ======================================================================


MAIN_PAGE_URL = "https://market.tase.co.il/he/market_data/securities/data/all"
API_URL = "https://api.tase.co.il/api/export/securitiesmarketdata"
BOI_RATES_URL = "https://boi.org.il/PublicApi/GetExchangeRates"

HEADERS = {
    "Accept": "text/csv",
    "Accept-Language": "he-IL",
    "Content-Type": "application/json;charset=UTF-8",
    "Origin": "https://market.tase.co.il",
    "Referer": "https://market.tase.co.il/",
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/149.0.0.0 Safari/537.36"
    ),
    "Sec-Ch-Ua": '"Google Chrome";v="149", "Chromium";v="149", "Not)A;Brand";v="24"',
    "Sec-Ch-Ua-Mobile": "?0",
    "Sec-Ch-Ua-Platform": '"Windows"',
    "Sec-Fetch-Dest": "empty",
    "Sec-Fetch-Mode": "cors",
    "Sec-Fetch-Site": "same-site",
}

PAYLOAD = {
    "FilterData": {"dType": 1, "TotalRec": 1, "pageNum": 1, "cl1": "0", "lang": "0"},
    "isAdd": False,
}

CSV_COL_NAME = 0
CSV_COL_TYPE = 3
CSV_COL_SEC_ID = 2
CSV_COL_PRICE = 4
CSV_COL_ATZMAT = 28
CSV_COL_MATURITY = 30
CSV_COL_CONV_STOCK_SEC_ID = 32
CSV_COL_YTM = 46
CSV_DATA_START_INDEX = 3

XLSX_SHEET_NAME = "Sheet1"
XLSX_COL_SEC_ID = "C"
XLSX_COL_PRICE = "E"
XLSX_COL_STOCK_PRICE = "AJ"


def log(msg: str, log_file):
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{timestamp}] {msg}"
    print(line)
    log_file.write(line + "\n")


def download_csv(log_file) -> pathlib.Path:
    DOWNLOADS_DIR.mkdir(exist_ok=True)
    session = requests.Session()

    log("מתחבר לדף הבורסה לקבלת cookies...", log_file)
    session.get(MAIN_PAGE_URL, headers={"User-Agent": HEADERS["User-Agent"]}, timeout=30)

    log("מוריד נתונים מה-API...", log_file)
    resp = session.post(API_URL, headers=HEADERS, json=PAYLOAD, timeout=60)

    if resp.status_code != 200 or "text/csv" not in resp.headers.get("Content-Type", ""):
        log(f"❌ ההורדה נכשלה - סטטוס {resp.status_code}", log_file)
        log(resp.text[:500], log_file)
        raise RuntimeError("הורדת ה-CSV נכשלה")

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M")
    csv_path = DOWNLOADS_DIR / f"tase_securities_{timestamp}.csv"
    csv_path.write_bytes(resp.content)
    log(f"✅ קובץ CSV נשמר: {csv_path}", log_file)
    return csv_path


def load_tase_csv(csv_path: pathlib.Path):
    with open(csv_path, encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))

    price_by_sec_id = {}
    conv_stock_sec_id_by_bond = {}

    for row in rows[CSV_DATA_START_INDEX:]:
        if len(row) <= max(CSV_COL_SEC_ID, CSV_COL_PRICE, CSV_COL_CONV_STOCK_SEC_ID):
            continue
        sec_id = row[CSV_COL_SEC_ID].strip()
        if not sec_id:
            continue
        price_raw = row[CSV_COL_PRICE].strip()
        if price_raw:
            try:
                price_by_sec_id[sec_id] = float(price_raw)
            except ValueError:
                pass
        conv_stock_id = row[CSV_COL_CONV_STOCK_SEC_ID].strip()
        if conv_stock_id:
            conv_stock_sec_id_by_bond[sec_id] = conv_stock_id.lstrip("0")

    return price_by_sec_id, conv_stock_sec_id_by_bond


def update_xlsx(csv_path: pathlib.Path, xlsx_path: pathlib.Path, log_file):
    price_by_sec_id, conv_stock_sec_id_by_bond = load_tase_csv(csv_path)
    log(f"נמצאו {len(price_by_sec_id)} ניירות ערך עם מחיר בקובץ CSV.", log_file)

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[XLSX_SHEET_NAME]

    updated_price = 0
    updated_stock_price = 0
    price_not_found = []
    stock_not_found = []

    for row_idx in range(2, ws.max_row + 1):
        sec_id_cell = ws[f"{XLSX_COL_SEC_ID}{row_idx}"]
        sec_id = str(sec_id_cell.value).strip() if sec_id_cell.value is not None else ""
        if not sec_id:
            continue

        if sec_id in price_by_sec_id:
            ws[f"{XLSX_COL_PRICE}{row_idx}"] = price_by_sec_id[sec_id]
            updated_price += 1
        else:
            price_not_found.append(sec_id)

        conv_stock_id = conv_stock_sec_id_by_bond.get(sec_id)
        if conv_stock_id and conv_stock_id in price_by_sec_id:
            ws[f"{XLSX_COL_STOCK_PRICE}{row_idx}"] = price_by_sec_id[conv_stock_id]
            updated_stock_price += 1
        elif conv_stock_id:
            stock_not_found.append(conv_stock_id)

    wb.save(xlsx_path)

    log(
        f"✅ עודכנו {updated_price} שערי אג\"ח, {updated_stock_price} שערי מניות המרה "
        f"(ערכים גולמיים, לפני חישוב מחדש)",
        log_file,
    )
    if price_not_found:
        log(f"⚠️  לא נמצאו {len(price_not_found)} מס' ני\"ע של אג\"ח: {price_not_found[:10]}", log_file)
    if stock_not_found:
        log(f"⚠️  לא נמצאו {len(stock_not_found)} מס' ני\"ע של מניות המרה: {stock_not_found[:10]}", log_file)


def recalculate_metrics(xlsx_path: pathlib.Path, log_file):
    log("מחשב מחדש YTM/פרמייה/Duration בפייתון (ללא Excel)...", log_file)
    recompute(str(xlsx_path))
    log("✅ החישוב הושלם ונשמר בקובץ.", log_file)


def fetch_usd_rate(log_file):
    """שולף שער יציג דולר/שקל עדכני מבנק ישראל. מחזיר (rate, date_str) או (None, None) בכשלון."""
    log("שולף שער דולר עדכני מבנק ישראל...", log_file)
    try:
        resp = requests.get(BOI_RATES_URL, timeout=20)
        resp.raise_for_status()
        data = resp.json()
        usd = next(r for r in data["exchangeRates"] if r["key"] == "USD")
        rate = usd["currentExchangeRate"]
        # lastUpdate: "2026-07-09T12:23:03.0575705Z" -> "09/07/2026"
        dt = datetime.datetime.fromisoformat(usd["lastUpdate"].replace("Z", "+00:00"))
        date_str = dt.strftime("%d/%m/%Y")
        log(f"✅ שער דולר: {rate} (מתאריך {date_str})", log_file)
        return rate, date_str
    except Exception as e:
        log(f"⚠️  שליפת שער הדולר נכשלה ({e}) - יישמר השער הקודם בדשבורד.", log_file)
        return None, None


def build_gov_curve(csv_path: pathlib.Path, log_file):
    """
    בונה עקום תשואות ממשלתי מתוך אג"ח 'ממשל שקלית' (לא צמודות) שבאותו קובץ
    CSV שכבר הורד - ללא צורך בשליפה חיצונית נוספת.
    """
    log("בונה עקום תשואות ממשלתי מתוך נתוני הבורסה שהורדו...", log_file)
    try:
        with open(csv_path, encoding="utf-8-sig") as f:
            rows = list(csv.reader(f))

        today = datetime.date.today()
        points = []
        for row in rows[CSV_DATA_START_INDEX:]:
            if len(row) <= CSV_COL_YTM:
                continue
            name = row[CSV_COL_NAME]
            if "ממשל שקלית" not in name:
                continue
            if row[CSV_COL_ATZMAT].strip() != "לא צמוד":
                continue
            maturity_raw = row[CSV_COL_MATURITY].strip()
            ytm_raw = row[CSV_COL_YTM].strip()
            if not maturity_raw or not ytm_raw:
                continue
            try:
                maturity_date = datetime.datetime.strptime(maturity_raw, "%d/%m/%Y").date()
                ytm = float(ytm_raw)
            except ValueError:
                continue
            t = (maturity_date - today).days / 365.25
            if t <= 0.05 or ytm == 0.0:
                continue  # קרוב מדי לפדיון / נתון חסר
            points.append({"t": round(t, 2), "y": ytm})

        points.sort(key=lambda p: p["t"])

        if len(points) < 3:
            log(f"⚠️  נמצאו רק {len(points)} נקודות לעקום - ייתכן שזה לא מספיק. נשמר העקום הקודם.", log_file)
            return None

        log(f"✅ נבנה עקום עם {len(points)} נקודות (טווח {points[0]['t']:.1f}-{points[-1]['t']:.1f} שנים).", log_file)
        return points
    except Exception as e:
        log(f"⚠️  בניית העקום הממשלתי נכשלה ({e}) - יישמר העקום הקודם בדשבורד.", log_file)
        return None


def build_dashboard(xlsx_path, dashboard_path, csv_path, usd_rate, usd_date, gov_curve, log_file):
    log("בונה מחדש את ytm_dashboard.html...", log_file)
    if not BUILD_DASHBOARD_SCRIPT.exists():
        log(f"❌ לא נמצא build_dashboard.py בנתיב: {BUILD_DASHBOARD_SCRIPT}", log_file)
        raise FileNotFoundError(BUILD_DASHBOARD_SCRIPT)

    cmd = [
        sys.executable,
        str(BUILD_DASHBOARD_SCRIPT),
        str(xlsx_path),
        "--template", str(dashboard_path),
        "--out", str(dashboard_path),
        "--nonconv-csv", str(csv_path),
    ]
    if usd_rate is not None:
        cmd += ["--usd-rate", str(usd_rate), "--usd-date", usd_date]
    if gov_curve is not None:
        cmd += ["--gov-curve", json.dumps(gov_curve)]
    if HOLDINGS_XLSX_PATH.exists():
        cmd += ["--holdings-xlsx", str(HOLDINGS_XLSX_PATH)]
    else:
        log(f"ℹ️  לא נמצא {HOLDINGS_XLSX_PATH.name} - לשוניות תיק ההחזקות לא יעודכנו.", log_file)
    if STOCK_FUNDAMENTALS_PATH.exists():
        cmd += ["--stock-fundamentals", str(STOCK_FUNDAMENTALS_PATH)]

    env = os.environ.copy()
    env["PYTHONIOENCODING"] = "utf-8"
    result = subprocess.run(
        cmd, capture_output=True, text=True, encoding="utf-8", errors="replace", env=env
    )
    log(result.stdout.strip(), log_file)
    if result.returncode != 0:
        log(f"❌ בניית הדשבורד נכשלה:\n{result.stderr}", log_file)
        raise RuntimeError("build_dashboard.py נכשל")
    if result.stderr.strip():
        log(f"(הודעות אזהרה מבניית הדשבורד):\n{result.stderr.strip()}", log_file)


def stamp_last_updated(dashboard_path: pathlib.Path, log_file):
    """מזריק שורת 'עודכן לאחרונה' (שעון ישראל) לראש קובץ ה-HTML הסופי."""
    now = datetime.datetime.now(ZoneInfo("Asia/Jerusalem"))
    stamp = now.strftime("%d/%m/%Y %H:%M")

    html = dashboard_path.read_text(encoding="utf-8")

    # מסירים חותמת קודמת אם קיימת (כדי לא לצבור כמה שורות בריצות חוזרות)
    html = re.sub(
        r'<div id="last-updated-banner"[^>]*>.*?</div>\s*',
        "",
        html,
        count=1,
        flags=re.DOTALL,
    )

    banner = (
        '<div id="last-updated-banner" '
        'style="background:#0f3460;color:#e8e8e8;text-align:center;'
        'padding:6px 8px;font-size:13px;font-family:Arial,sans-serif;'
        'direction:rtl;">'
        f"עודכן לאחרונה: {stamp} (שעון ישראל)"
        "</div>\n"
    )

    new_html, n = re.subn(r"(<body[^>]*>)", r"\1\n" + banner, html, count=1)
    if n == 0:
        log("⚠️ לא נמצא תג <body> בדשבורד - החותמת לא נוספה.", log_file)
        return

    dashboard_path.write_text(new_html, encoding="utf-8")
    log(f"✅ נוספה חותמת 'עודכן לאחרונה: {stamp}' לדשבורד.", log_file)


def main():
    xlsx_path = pathlib.Path(XLSX_PATH)
    dashboard_path = pathlib.Path(DASHBOARD_PATH)

    with open(LOG_PATH, "a", encoding="utf-8") as log_file:
        log_file.write("\n" + "=" * 60 + "\n")
        log("--- תחילת ריצה ---", log_file)

        if not xlsx_path.exists():
            log(f"❌ קובץ ה-Excel לא נמצא בנתיב: {xlsx_path}", log_file)
            sys.exit(1)
        if not dashboard_path.exists():
            log(f"❌ קובץ הדשבורד לא נמצא בנתיב: {dashboard_path}", log_file)
            sys.exit(1)

        try:
            csv_path = download_csv(log_file)

            new_bonds = add_new_bonds(str(csv_path), str(xlsx_path), log_func=lambda m: log(m, log_file))
            if new_bonds:
                log(f"זוהו {len(new_bonds)} אג\"ח להמרה חדשות ונוספו אוטומטית: {', '.join(new_bonds)}", log_file)

            update_xlsx(csv_path, xlsx_path, log_file)
            recalculate_metrics(xlsx_path, log_file)
            usd_rate, usd_date = fetch_usd_rate(log_file)
            gov_curve = build_gov_curve(csv_path, log_file)
            build_dashboard(xlsx_path, dashboard_path, csv_path, usd_rate, usd_date, gov_curve, log_file)
            stamp_last_updated(dashboard_path, log_file)
            log("--- ריצה הסתיימה בהצלחה (כולל דולר, עקום ודשבורד) ---", log_file)
        except Exception as e:
            log(f"❌ שגיאה כללית: {e}", log_file)
            sys.exit(1)


if __name__ == "__main__":
    main()
