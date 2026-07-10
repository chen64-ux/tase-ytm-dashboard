# -*- coding: utf-8 -*-
"""
add_new_convertible_bonds.py
מזהה אג"ח להמרה חדשות בקובץ ה-CSV שהורד מהבורסה (שעדיין לא קיימות ב-
ytm_computed.xlsx) ומוסיף אותן כשורה חדשה, עם כל השדות שהבורסה מפרסמת.

תנאי זיהוי (כל 4 חייבים להתקיים בשורת ה-CSV):
    1. סוג הנייר = "איגרות חוב להמרה"
    2. יש מספר ני"ע למניית ההמרה (עמודה AG)
    3. יש יחס המרה - כמות אג"ח למניה אחת (עמודה AI)
    4. יש מועד המרה סופי (עמודה AK)

שימוש (כמודול, מתוך run_daily_update.py):
    from add_new_convertible_bonds import add_new_bonds
    added = add_new_bonds(csv_path, xlsx_path, log_func=log)

הרצה עצמאית לבדיקה:
    python3 add_new_convertible_bonds.py <csv_path> <xlsx_path>
"""

import csv
import sys
from datetime import datetime

import openpyxl

# מיפוי: אינדקס עמודה ב-CSV (0-indexed) -> אות עמודה ב-Excel
# (נבדק מול securitiesmarketdata.csv האמיתי - היסט קבוע של +1 בין השניים)
CSV_TO_XLSX = {
    0: "A",    # שם
    1: "B",    # סימול
    2: "C",    # מס' ני"ע
    3: "D",    # סוג ני"ע
    4: "E",    # שער אחרון
    28: "AC",  # הצמדה
    29: "AD",  # ריבית
    30: "AE",  # מועד פדיון
    31: "AF",  # מניית המרה (שם)
    32: "AG",  # מניית המרה - מס' ני"ע
    33: "AH",  # מניית המרה - ISIN
    34: "AI",  # כמות אג"ח למניה אחת (יחס המרה)
    35: "AJ",  # שער מניית המרה
    36: "AK",  # מועד המרה סופי
    48: "AW",  # ISIN
    49: "AX",  # מספר מנפיק
}

DATE_COLS = {30, 36}  # אינדקסים ב-CSV שהם תאריכים (מועד פדיון, מועד המרה סופי)
NUM_COLS = {4, 29, 34, 35}  # שדות מספריים (שער, ריבית, יחס המרה, שער מניה)

CONVERTIBLE_TYPE = "איגרות חוב להמרה"


def _parse_date(s: str):
    s = s.strip()
    if not s:
        return None
    return datetime.strptime(s, "%d/%m/%Y")


def _parse_num(s: str):
    s = s.strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _read_csv_convertibles(csv_path: str) -> dict:
    """מחזיר dict: sec_id -> שורת CSV גולמית, רק עבור אג"ח להמרה עם כל השדות הדרושים."""
    with open(csv_path, encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))
    data = rows[3:]  # 3 שורות כותרת בראש הקובץ (כותרת כללית, תאריך, שמות עמודות)

    result = {}
    for row in data:
        if len(row) <= 36:
            continue
        if row[3].strip() != CONVERTIBLE_TYPE:
            continue
        if not (row[32].strip() and row[34].strip() and row[36].strip()):
            continue
        sec_id = row[2].strip()
        if sec_id:
            result[sec_id] = row
    return result


def _existing_sec_ids(ws) -> set:
    ids = set()
    for r in range(2, ws.max_row + 1):
        v = ws[f"C{r}"].value
        if v:
            ids.add(str(v).strip())
    return ids


def add_new_bonds(csv_path: str, xlsx_path: str, log_func=print) -> list:
    """
    מוסיף לקובץ ה-Excel שורות חדשות עבור אג"ח להמרה שנמצאות ב-CSV
    ועדיין לא נמצאות בקובץ. מחזיר רשימת שמות האג"ח שנוספו (ריקה אם אין חדשות).
    """
    csv_convertibles = _read_csv_convertibles(csv_path)

    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    ws = wb["Sheet1"]
    existing_ids = _existing_sec_ids(ws)

    new_ids = [sid for sid in csv_convertibles if sid not in existing_ids]
    if not new_ids:
        return []

    added_names = []
    next_row = ws.max_row + 1
    for sid in new_ids:
        row = csv_convertibles[sid]
        for csv_idx, xlsx_col in CSV_TO_XLSX.items():
            if csv_idx >= len(row):
                continue
            raw = row[csv_idx]
            if csv_idx in DATE_COLS:
                value = _parse_date(raw)
            elif csv_idx in NUM_COLS:
                value = _parse_num(raw)
            else:
                value = raw.strip() if raw else None
            ws[f"{xlsx_col}{next_row}"] = value

        name = row[0].strip()
        added_names.append(name)
        log_func(f"  ➕ אג\"ח להמרה חדשה זוהתה ונוספה: {name} (מס' ני\"ע {sid})")
        next_row += 1

    wb.save(xlsx_path)
    log_func(
        f"⚠️  נוספו {len(added_names)} אג\"ח להמרה חדשות ל-{xlsx_path}. "
        f"מומלץ לפתוח ולבדוק חזותית פעם אחת שהשורות תקינות."
    )
    return added_names


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("שימוש: python3 add_new_convertible_bonds.py <csv_path> <xlsx_path>")
        sys.exit(1)
    added = add_new_bonds(sys.argv[1], sys.argv[2])
    if added:
        print(f"נוספו {len(added)} אג\"ח: {added}")
    else:
        print("לא נמצאו אג\"ח להמרה חדשות.")
